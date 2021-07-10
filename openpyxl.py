"""
このtemplate.pyの中身をコピーして
サイト構築をしていこう!!

streamlit_web.py ファイルを開いて、
1. sidebar_list リストにファイル名を追加する
2. import で自作ファイルを読み出す

"""

# 最初に辞書型を定義
index_dict = {}

# ******このエリアにindex_dict[] = を書いていく********************

# openpyxl のインポート
index_dict['openpyxl のインポート'] = (
    """
1. オーソドックスなimport (全てのopenpyxlを呼び出すイメージ)
```python
import openpyxl
```
2. 既存のファイルを読み込むだけならこれでも!!
- openpyxl内のload_workbookのみをインポート
```python
from openpyxl import load_workbook
```
    """
)

# Excelファイルの読込みしワークブックを指定 -> シート を読込む
index_dict['Excelファイルの読込み -> ワークブックを指定 -> シート を読込む'] = (
    """
## openpyxlのインポートの仕方で少し変わる。

#### Excelファイルの読込み -> ワークブックを指定

1. import openpyxl でインポートした場合
```python
wb = openpyxl.load_workbook(filename='※ここにExcelファイル※.xlsx')
```

2. from openpyxl import load_workbook でインポートした場合

```python
wb = load_workbook(filename='※ここにExcelファイル※.xlsx')
```

#### 指定したワークブック -> シートを読込む

1. sheetnames を使用することでシート名をリスト型で取得する。
```python
sheet_name_list = wb.sheetnames
print(sheet_name_list, type(sheet_name_list))
```

2. シート名を指定してシートの内容を取得する。(例:シートの先頭を指定)
```python
sheet_name = wb.sheetnames[0]
ws = wb[sheet_name]
```

    """
)

# ワークブックに新しいシートの追加 (create_sheet)
index_dict['ワークブックに新しいシートの追加 (create_sheet)'] = (
    """
1. create_sheet() メソッドを使う。

```python
wb = load_workbook(filename='※ここにExcelファイル※.xlsx')
new_ws = wb.create_sheet('※新しく作るシート名を入れる※')
```
    """
)

# 編集したワークブックを名前を変えて保存する。 (save)
index_dict['編集したワークブックを名前を変えて保存する。 (save)'] = (
    """
1. save() メソッドを使う。

```python
wb = load_workbook(filename='※ここにExcelファイル※.xlsx')
sheet_name = wb.sheetnames[0]
ws = wb[sheet_name]

# NOTE: ここで色々編集があったとしよう!!

wb.save('※ここに保存したいExcelファイル名を入力する※.xlsx')

```
    """
)

# よく使うExcel操作(セルの結合)
index_dict['よく使うExcel操作(セルの結合)'] = (
    """
1. セルの結合 merge_cells() メソッドを使う。
- ()の引数はセルの範囲を指定する。
- 結合したセルに代入する時は、一番左上を指定する。
```python
ws.merge_cells('A6:K10')
```
- 指定している場所の図
![](https://user-images.githubusercontent.com/79512367/124370642-daeb2680-dcb4-11eb-9bae-32131c4b7104.png)

- 参考図
![](https://user-images.githubusercontent.com/79512367/124370292-b3925a80-dcb0-11eb-99e1-40bd554b6551.png)

    """
)

# ************************************************************

# index_listのキーデータをリストにする。
index_keys_list = list(index_dict.keys())


# index_send関数
def index_send():
    """
    この関数を呼び出すと、
    このファイル内の出力可能なindex_listを返してくる。
    """
    return index_keys_list


# text_send関数
def text_send(index_keys_set):
    """
    この関数が呼び出されると、
    pythonコードのマークダウンが出力される!!
    """
    return index_dict[index_keys_set]


if __name__ == '__main__':
    index_send()
    # text_send('type() 型を確認する')
